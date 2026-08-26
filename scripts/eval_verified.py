"""Score a model against every human-verified mask, per subject and per image.

Unlike the training val_dice (a random 20% split of training images), this
evaluates against ALL verified masks — and once holdout subjects have verified
labels, it reports them separately as the honest generalization number.

Usage:
    PythonSlicer scripts/eval_verified.py                     # active model
    PythonSlicer scripts/eval_verified.py --model models/ucl_seg_v8.pt
    PythonSlicer scripts/eval_verified.py --overlays eval_out/   # side-by-side PNGs
    PythonSlicer scripts/eval_verified.py --log     # append to training_runs.xlsx 'Eval History'
"""
from __future__ import annotations
import argparse, json, sys
from pathlib import Path
import numpy as np
import torch
from PIL import Image

PIPELINE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PIPELINE))
from ucl.model import UNet
from ucl.data import load_mask, pad_to_multiple, unpad, SEG_CLASS_MAP

ID_TO_NAME = {v: k for k, v in SEG_CLASS_MAP.items()}


def read_gray(fp):
    if fp.suffix.lower() == ".dcm":
        import pydicom
        arr = pydicom.dcmread(str(fp)).pixel_array
        if arr.ndim == 3:
            arr = arr.mean(axis=2)
        if arr.ndim != 2:
            raise ValueError(f"unsupported pixel array shape {arr.shape}")
        return arr.astype(np.uint8)
    return np.asarray(Image.open(fp).convert("L"))


@torch.no_grad()
def predict(model, gray, device, resize):
    H0, W0 = gray.shape
    img = gray.astype(np.float32) / 255.0
    if resize:
        rh, rw = resize
        img = np.asarray(Image.fromarray((img*255).astype(np.uint8))
                         .resize((rw, rh), Image.BILINEAR), np.float32) / 255.0
    img_p, pad = pad_to_multiple(img, 16)
    t = torch.from_numpy(img_p)[None, None].to(device)
    labels = torch.softmax(model(t), 1)[0].argmax(0).cpu().numpy().astype(np.int32)
    labels = unpad(labels, pad)
    if labels.shape != (H0, W0):
        labels = np.asarray(Image.fromarray(labels.astype(np.uint8))
                            .resize((W0, H0), Image.NEAREST), np.int32)
    return labels


def dice(p, t):
    denom = p.sum() + t.sum()
    return 2 * np.logical_and(p, t).sum() / denom if denom else float("nan")


def save_overlay_pair(gray, gt, pred, scores, title, out_path):
    import cv2
    colors = {1: (0, 153, 255), 2: (0, 255, 0)}  # BGR
    def paint(mask):
        rgb = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
        for cid, col in colors.items():
            m = (mask == cid).astype(np.uint8)
            if not m.any(): continue
            tint = rgb.copy(); tint[m > 0] = col
            rgb = cv2.addWeighted(rgb, 0.65, tint, 0.35, 0)
            cnts, _ = cv2.findContours(m, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(rgb, cnts, -1, col, 2)
        return rgb
    left, right = paint(gt), paint(pred)
    cv2.putText(left,  "HUMAN (ground truth)", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255,255,255), 2)
    cv2.putText(right, "MODEL prediction",     (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255,255,255), 2)
    panel = np.hstack([left, np.full((gray.shape[0], 8, 3), 255, np.uint8), right])
    bar = np.zeros((44, panel.shape[1], 3), np.uint8)
    cv2.putText(bar, f"{title}   " + "   ".join(f"{ID_TO_NAME[c]} {s:.3f}" for c, s in scores.items()),
                (12, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255,255,255), 2)
    cv2.imwrite(str(out_path), np.vstack([bar, panel]))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--model",    default=str(PIPELINE / "models" / "ucl_seg.pt"))
    ap.add_argument("--overlays", default=None, help="directory for side-by-side PNGs")
    ap.add_argument("--log",      action="store_true",
                    help="append per-subject means to training_runs.xlsx 'Eval History'")
    args = ap.parse_args()

    device = "cuda" if torch.cuda.is_available() else "cpu"
    ck = torch.load(args.model, map_location=device)
    nc = ck.get("num_classes", 3)
    model = UNet(in_ch=1, out_ch=nc, base=ck.get("base", 32)).to(device)
    model.load_state_dict(ck["model"]); model.eval()
    resize = ck.get("resize")
    print(f"model: {Path(args.model).name}  classes={nc}  resize={resize}  device={device}\n")

    holdout = set()
    hf = PIPELINE / "holdout_subjects.json"
    if hf.exists():
        holdout = set(json.loads(hf.read_text()).get("holdout_subjects", []))

    out_dir = Path(args.overlays) if args.overlays else None
    if out_dir: out_dir.mkdir(parents=True, exist_ok=True)

    class_ids = sorted(ID_TO_NAME)
    rows = []          # (subject, stem, {cid: dice})
    for subj in sorted((PIPELINE / "subjects").iterdir()):
        if not subj.is_dir(): continue
        for sess in sorted((subj / "sessions").glob("*")):
            mask_dir, img_dir = sess / "masks", sess / "images"
            if not mask_dir.exists(): continue
            for mp in sorted(mask_dir.glob("*.nii.gz")):
                stem = mp.name.replace(".nii.gz", "")
                fp = next((img_dir / (stem + e) for e in (".dcm", ".png", ".jpg")
                           if (img_dir / (stem + e)).exists()), None)
                if fp is None: continue
                try:
                    gray = read_gray(fp)
                    gt   = load_mask(stem, mask_dir, nc)
                    if gt.shape != gray.shape:
                        print(f"  !! shape mismatch {subj.name}/{stem}"); continue
                    pred = predict(model, gray, device, resize)
                except Exception as e:
                    print(f"  !! {subj.name}/{stem}: {e}"); continue
                d = {c: dice(pred == c, gt == c) for c in class_ids}
                rows.append((subj.name, stem, d))
                if out_dir:
                    save_overlay_pair(gray, gt, pred, d, f"{subj.name}/{stem}",
                                      out_dir / f"{subj.name}__{stem}.png")

    if not rows:
        raise SystemExit("No verified masks with matching images found.")

    hdr = f"{'subject':<16}{'image':<34}" + "".join(f"{ID_TO_NAME[c]:>10}" for c in class_ids) + f"{'mean':>9}"
    print(hdr)
    for subj, stem, d in rows:
        vals = [d[c] for c in class_ids]
        tag  = " (holdout)" if subj in holdout else ""
        print(f"{subj+tag:<16}{stem[:32]:<34}" + "".join(f"{v:>10.3f}" for v in vals)
              + f"{np.nanmean(vals):>9.3f}")

    print("-" * len(hdr))
    subject_means = {}
    for subj in sorted({r[0] for r in rows}):
        sub = [r[2] for r in rows if r[0] == subj]
        means = {c: float(np.nanmean([d[c] for d in sub])) for c in class_ids}
        subject_means[subj] = (len(sub), means)
        tag = " (holdout)" if subj in holdout else ""
        print(f"{subj+tag+' mean':<50}" + "".join(f"{means[c]:>10.3f}" for c in class_ids)
              + f"{np.nanmean(list(means.values())):>9.3f}")
    overall = {c: float(np.nanmean([r[2][c] for r in rows])) for c in class_ids}
    print(f"{'OVERALL ('+str(len(rows))+' images)':<50}"
          + "".join(f"{overall[c]:>10.3f}" for c in class_ids)
          + f"{np.nanmean(list(overall.values())):>9.3f}")

    if args.log:
        log_eval(args.model, subject_means, overall, len(rows), holdout, class_ids)
    if out_dir:
        print(f"\noverlays → {out_dir}")


def log_eval(model_path, subject_means, overall, n_images, holdout, class_ids):
    import datetime
    header = ["Date", "Model", "Subject", "Images"] + \
             [ID_TO_NAME[c].capitalize() + " Dice" for c in class_ids] + ["Mean Dice", "Holdout?"]
    date, mname = datetime.date.today().isoformat(), Path(model_path).name
    out_rows = []
    for subj, (n, means) in subject_means.items():
        vals = [round(means[c], 4) for c in class_ids]
        out_rows.append([date, mname, subj, n] + vals +
                        [round(float(np.nanmean(vals)), 4), "YES" if subj in holdout else "No"])
    ov = [round(overall[c], 4) for c in class_ids]
    out_rows.append([date, mname, "OVERALL", n_images] + ov +
                    [round(float(np.nanmean(ov)), 4), ""])
    xlsx = PIPELINE / "training_runs.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.load_workbook(str(xlsx))
        ws = wb["Eval History"] if "Eval History" in wb.sheetnames else wb.create_sheet("Eval History")
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            ws.append(header)
            for c in ws[1]: c.font = Font(name="Arial", size=11, bold=True)
        for row in out_rows:
            ws.append(row)
            for c in ws[ws.max_row]: c.font = Font(name="Arial", size=11)
        wb.save(str(xlsx))
        print(f"\neval logged → {xlsx.name} (Eval History)")
    except Exception as e:
        print(f"\ncould not log to xlsx ({e}) — results above are printed only")


if __name__ == "__main__":
    main()
