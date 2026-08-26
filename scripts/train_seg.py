"""
Train the UCL segmentation U-Net.

Near-verbatim copy of scripts/train.py from the patellar reference.
Only differences:
  - uses UCLSegDataset instead of TendonSegDataset
  - default --num_classes 4 (bg + ucl + bone_humerus + bone_ulna)
  - checkpoint tagged with task="segmentation"

Usage:
    python scripts/train_seg.py --data _train_seg/ --epochs 80 \
        --resize 320 512 --out models/ucl_seg.pt
"""
from __future__ import annotations
import argparse
from pathlib import Path
import numpy as np
import torch
from torch.utils.data import DataLoader, random_split
import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ucl.model import UNet, dice_ce_loss
from ucl.data  import UCLSegDataset, NUM_SEG_CLASSES


@torch.no_grad()
def val_dice(model, loader, device, nc):
    """Mean foreground Dice — identical to train.py val_dice."""
    model.eval(); per_class = []
    for img, msk in loader:
        img, msk = img.to(device), msk.to(device)
        pred = model(img).argmax(1)
        dices = []
        for c in range(1, nc):
            p = (pred == c).float(); t = (msk == c).float()
            if t.sum() > 0:
                dices.append(((2*(p*t).sum()+1e-6)/(p.sum()+t.sum()+1e-6)).item())
        if dices: per_class.append(np.mean(dices))
    return float(np.mean(per_class)) if per_class else 0.0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data",        default=str(Path(__file__).resolve().parents[1]/"_train_seg"),
                    help="training data root (auto-built from subjects/ if not specified)")
    ap.add_argument("--epochs",      type=int,   default=80)
    ap.add_argument("--batch",       type=int,   default=4)
    ap.add_argument("--lr",          type=float, default=1e-3)
    ap.add_argument("--base",        type=int,   default=32)
    ap.add_argument("--num_classes", type=int,   default=NUM_SEG_CLASSES)
    ap.add_argument("--val_frac",    type=float, default=0.2)
    ap.add_argument("--resize",      type=int,   nargs=2, default=None,
                    metavar=("H", "W"))
    ap.add_argument("--out",         default="models/ucl_seg.pt")
    ap.add_argument("--seed",        type=int,   default=0)
    ap.add_argument("--backbone",    default="unet", choices=["unet","ultrasam"],
                    help="unet = train from scratch (default). "
                         "ultrasam = UltraSam ViT encoder + U-Net decoder.")
    ap.add_argument("--ultrasam_ckpt", default="models/UltraSam.pth",
                    help="path to UltraSam.pth (only with --backbone ultrasam)")
    ap.add_argument("--freeze_epochs", type=int, default=20,
                    help="epochs to freeze UltraSam encoder before full fine-tune")
    args = ap.parse_args()

    torch.manual_seed(args.seed); np.random.seed(args.seed)

    # auto-build training set from subjects/ folder
    import importlib.util
    bts_path = Path(__file__).resolve().parent / "build_training_set.py"
    if bts_path.exists():
        spec = importlib.util.spec_from_file_location("build_training_set", bts_path)
        bts  = importlib.util.module_from_spec(spec); spec.loader.exec_module(bts)
        n = bts.main()
        if not n:
            raise SystemExit("No labeled images found. Label images in Slicer first.")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"device: {device}")

    resize = tuple(args.resize) if args.resize else None
    nc     = args.num_classes
    full_tr = UCLSegDataset(args.data, augment=True,  resize=resize, num_classes=nc)
    full_va = UCLSegDataset(args.data, augment=False, resize=resize, num_classes=nc)

    n = len(full_tr); n_val = max(1, int(n * args.val_frac)); n_tr = n - n_val
    g = torch.Generator().manual_seed(args.seed)
    ti, vi = random_split(range(n), [n_tr, n_val], generator=g)
    tr = torch.utils.data.Subset(full_tr, list(ti))
    va = torch.utils.data.Subset(full_va, list(vi))
    print(f"train {len(tr)}  val {len(va)}  classes {nc}")

    tl = DataLoader(tr, batch_size=args.batch, shuffle=True,  num_workers=0)
    vl = DataLoader(va, batch_size=1,          shuffle=False, num_workers=0)

    # ---- model selection ----
    if args.backbone == "ultrasam":
        from ucl.model import UltraSamUNet
        ckpt_path = Path(args.ultrasam_ckpt)
        if not ckpt_path.exists():
            raise SystemExit(
                f"UltraSam checkpoint not found: {ckpt_path}\n"
                f"Download it with:\n"
                f"  curl -L -o {ckpt_path} "
                f"https://s3.unistra.fr/camma_public/github/ultrasam/UltraSam.pth"
            )
        print(f"Using UltraSam backbone from {ckpt_path}")
        print(f"  Phase A: encoder frozen for {args.freeze_epochs} epochs")
        print(f"  Phase B: full fine-tune for {args.epochs - args.freeze_epochs} epochs")
        model = UltraSamUNet(checkpoint=str(ckpt_path), out_ch=nc,
                             freeze_encoder=True).to(device)
        backbone_tag = "ultrasam"
    else:
        model = UNet(in_ch=1, out_ch=nc, base=args.base).to(device)
        backbone_tag = "unet"

    opt   = torch.optim.Adam(
        filter(lambda p: p.requires_grad, model.parameters()), lr=args.lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=args.epochs)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    best = -1.0
    for ep in range(1, args.epochs+1):

        # Phase B: unfreeze encoder after freeze_epochs
        if args.backbone == "ultrasam" and ep == args.freeze_epochs + 1:
            print(f"\n--- Phase B: unfreezing UltraSam encoder ---")
            model.unfreeze_encoder()
            # rebuild optimizer to include encoder params
            opt = torch.optim.Adam(model.parameters(), lr=args.lr * 0.1)
            sched = torch.optim.lr_scheduler.CosineAnnealingLR(
                opt, T_max=args.epochs - args.freeze_epochs)

        model.train(); losses = []
        for img, msk in tl:
            img, msk = img.to(device), msk.to(device)
            opt.zero_grad()
            loss = dice_ce_loss(model(img), msk, num_classes=nc)
            loss.backward(); opt.step()
            losses.append(loss.item())
        sched.step()
        d = val_dice(model, vl, device, nc)
        phase = "A" if (args.backbone=="ultrasam" and ep<=args.freeze_epochs) else "B"
        print(f"epoch {ep:3d} [{phase}]  loss {np.mean(losses):.4f}  val_dice {d:.4f}")
        if d > best:
            best = d
            torch.save({"model": model.state_dict(), "base": args.base,
                        "num_classes": nc, "resize": resize,
                        "val_dice": best, "task": "segmentation",
                        "backbone": backbone_tag}, args.out)
            print(f"  saved → {args.out}  (dice {best:.4f})")
    print(f"\ndone. best val dice {best:.4f}")
    log_run(args, resize, len(tr), len(va), best, backbone_tag)


def log_run(args, resize, n_tr, n_val, best, backbone_tag):
    """Append this run to training_runs.xlsx (sheet 'Training Runs');
    fall back to training_runs.csv if openpyxl isn't installed."""
    import datetime, json, platform
    root = Path(__file__).resolve().parents[1]
    n_labels = len(list((root / "subjects").glob("*/sessions/*/masks/*.nii.gz")))
    note = ""
    try:
        prov = json.loads((root / "label_provenance.json").read_text())
        n_corr = sum(1 for v in prov.values() if v.get("source") == "corrected_prelabel")
        note = f"{n_labels - n_corr} baseline + {n_corr} corrected pre-labels"
    except Exception: pass
    row = [Path(args.out).stem, datetime.date.today().isoformat(),
           platform.node(),
           "cuda" if torch.cuda.is_available() else "cpu",
           backbone_tag, args.epochs,
           f"{resize[0]}x{resize[1]}" if resize else "none",
           args.batch, args.lr, n_tr, n_val, n_labels,
           round(best, 4), Path(args.out).name, note]
    xlsx = root / "training_runs.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx))
        ws = wb["Training Runs"]
        ws.append(row)
        from openpyxl.styles import Font
        for c in ws[ws.max_row]: c.font = Font(name="Arial", size=11)
        wb.save(str(xlsx))
        print(f"run logged → {xlsx.name}")
    except Exception as e:
        import csv
        csvp = root / "training_runs.csv"
        new = not csvp.exists()
        with open(csvp, "a", newline="") as f:
            w = csv.writer(f)
            if new:
                w.writerow(["Version","Date","Machine","Device","Backbone","Epochs",
                            "Resize","Batch","LR","Train Imgs","Val Imgs",
                            "Verified Labels","Val Dice","Model File","Notes"])
            w.writerow(row)
        print(f"run logged → {csvp.name} (xlsx unavailable: {e}) — "
              f"paste this row into training_runs.xlsx when convenient")


if __name__ == "__main__":
    main()
